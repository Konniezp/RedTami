import base64
from datetime import datetime, date
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')
from io import BytesIO
import requests
import numpy as np

from django.contrib import auth, messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.db import connection
from django.db.models import Count, F, Max
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.utils import timezone


from openpyxl import Workbook

from rest_framework.views import APIView
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import viewsets
from rest_framework.authentication import SessionAuthentication, BasicAuthentication
from rest_framework.permissions import IsAdminUser

from .forms import *
from .models import *
from .serializer import *



    
@login_required
def home(request):
    return render(request, "home.html")


def login(request):
    if request.method == "POST":
        username = request.POST["username"]
        password = request.POST["password"]

        user = auth.authenticate(username=username, password=password)

        if user is not None:
            auth.login(request, user)
            return redirect("home")
        else:
            return render(request, "registration/login.html")

    return render(request, "registration/login.html")

# --------------------- Respuestas de Usuario --------------------- #

#Home
@login_required
def respuestasHome(request):
    return render(request, "respuestas/respuestasHome.html")


# Base de datos
@login_required
def datosPerfil(request):
    Datos = Usuario.objects.all().order_by("-Fecha_Ingreso")
    data = {
        "Datos": Datos,
    }
    return render(request, "respuestas/datosPerfil.html", data)

@login_required
def datosPreguntas(request):
    Datos = UsuarioRespuesta.objects.all().order_by("-fecha_respuesta")
    data = {
        "Datos": Datos,
    }
    return render(request, "respuestas/datosPreguntas.html", data)

@login_required
def datosTextoPreguntas(request):
    Datos = UsuarioTextoPregunta.objects.all().order_by("-fecha_pregunta")
    data = {
        "Datos": Datos,
    }
    return render(request, "respuestas/datosPreguntasEspecialistas.html", data)

def crear_excel_desde_db():
    # Crear un nuevo libro de trabajo de Excel
    wb = Workbook()

    # Hoja para respuestas de usuarios
    ws_respuestas_usuario = wb.active
    ws_respuestas_usuario.title = 'Respuestas Usuario'

    # Obtener todas las preguntas y almacenarlas en una lista
    preguntas = Pregunta.objects.all()
    lista_preguntas = ['Rut'] + [pregunta.pregunta for pregunta in preguntas]

    # Agregar las preguntas a la primera fila del archivo Excel
    ws_respuestas_usuario.append(lista_preguntas)

    # Obtener los datos de los usuarios y sus respuestas
    usuarios_respuestas = UsuarioRespuesta.objects.select_related('id_opc_respuesta', 'id_opc_respuesta__id_pregunta').values(
        'Rut',
        'id_opc_respuesta__id_pregunta__pregunta',
        'id_opc_respuesta__OPC_Respuesta'
    )

    # Crear un diccionario para almacenar las respuestas de los usuarios
    dict_respuestas = {}

    # Agregar las respuestas de los usuarios al diccionario
    for respuesta in usuarios_respuestas:
        rut = respuesta['Rut']
        pregunta = respuesta['id_opc_respuesta__id_pregunta__pregunta']
        respuesta_usuario = respuesta['id_opc_respuesta__OPC_Respuesta']
        if rut not in dict_respuestas:
            dict_respuestas[rut] = {}
        dict_respuestas[rut][pregunta] = respuesta_usuario

    # Agregar las respuestas de los usuarios al archivo Excel
    for rut, respuestas_usuario in dict_respuestas.items():
        fila = [rut]
        for pregunta in preguntas:
            respuesta = respuestas_usuario.get(pregunta.pregunta, '')
            fila.append(respuesta)
        ws_respuestas_usuario.append(fila)

    # Hoja para datos del perfil de usuario
    ws_datos_perfil = wb.create_sheet(title='Datos Perfil')

    # Obtener los nombres de los campos del modelo Usuario
    campos_usuario = [field.name for field in Usuario._meta.fields if field.name not in ['Comuna_Usuario', 'Genero_Usuario', 'SistemaSalud_Usuario', 'Ocupacion_Usuario']]

    # Agregar los nombres de los campos a la primera fila del archivo Excel
    ws_datos_perfil.append(campos_usuario)

    # Obtener los datos de los usuarios y agregarlos al archivo Excel
    for usuario in Usuario.objects.all():
        # Convertir la fecha a formato de texto para evitar problemas con zonas horarias
        datos_usuario = [str(getattr(usuario, campo)) for campo in campos_usuario]
        ws_datos_perfil.append(datos_usuario)

    # Hoja para preguntas al especialista
    ws_preguntas_especialista = wb.create_sheet(title='Preguntas al especialista')

    # Obtener los nombres de los campos del modelo UsuarioTextoPregunta
    campos_preguntas_especialista = [field.name for field in UsuarioTextoPregunta._meta.fields if field.name != 'id']

    # Agregar los nombres de los campos a la primera fila del archivo Excel
    ws_preguntas_especialista.append(campos_preguntas_especialista)

    # Obtener los datos de las preguntas al especialista y agregarlos al archivo Excel
    for pregunta in UsuarioTextoPregunta.objects.all():
        # Convertir la fecha a formato de texto para evitar problemas con zonas horarias
        datos_pregunta = [str(getattr(pregunta, campo)) for campo in campos_preguntas_especialista]
        ws_preguntas_especialista.append(datos_pregunta)

    # Guardar el libro de trabajo en un archivo
    nombre_archivo = 'reporte_respuestas.xlsx'
    wb.save(nombre_archivo)

    return nombre_archivo

def descargar_excel(request):
    # Llama a la función para crear el Excel
    nombre_archivo = crear_excel_desde_db()

    # Abre el archivo Excel y lo envía como una respuesta de descarga
    with open(nombre_archivo, 'rb') as excel_file:
        response = HttpResponse(excel_file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={nombre_archivo}'
        return response
# --------------------- Reporteria --------------------- #

# Configuración global para fuentes de gráficos
plt.rcParams['font.family'] = 'sans-serif'  
plt.rcParams['font.sans-serif'] = 'Calibri' 
plt.rcParams['font.size'] = 11
plt.rcParams['axes.titlesize'] =20
plt.rcParams['axes.labelsize']= 13
plt.rcParams['axes.labelpad']=10

def generar_grafico_usuario_por_edad():

    with connection.cursor() as cursor:
        cursor.execute(
            "SELECT edad, COUNT(*) FROM botApp_usuario GROUP BY edad ORDER BY edad "
        )
        resultados = cursor.fetchall()

    edades = []
    cantidades = []

    for resultado in resultados:
        edad, cantidad = resultado
        edades.append(edad)
        cantidades.append(cantidad)

    plt.figure(figsize=[13,5])
    plt.bar(edades, cantidades, color="blue")
    plt.xlabel("Edad")
    plt.ylabel("Número de Usuarias")
    plt.title("Usuarias por edad", pad=20)
    plt.xticks(range(min(edades), max(edades) + 1, 1))
    

    # Agregar etiquetas en las barras
    for edad, cantidad in zip(edades, cantidades):
        plt.text(edad, cantidad, str(cantidad), ha='center', va='bottom')

    # Guardar la imagen en un buffer
    buffer = BytesIO()
    plt.savefig(buffer, format="png")
    buffer.seek(0)
    plt.close()

    # Convertir la imagen a base64
    imagen_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
    return imagen_base64


def generar_grafico_anio_nacimiento():
    with connection.cursor() as cursor:
        cursor.execute(
            "SELECT YEAR(AnioNacimiento) as anio, COUNT(*) FROM botApp_usuario GROUP BY YEAR(AnioNacimiento) ORDER BY anio ASC;"
        )
        resultados = cursor.fetchall()

    anios = []
    cantidades = []

    for resultado in resultados:
        anio, cantidad = resultado
        anios.append(anio)
        cantidades.append(cantidad)

    plt.figure(figsize=[10,7])
    plt.bar(anios, cantidades, color="blue")
    plt.xlabel("Año de Nacimiento")
    plt.ylabel("Número de Usuarios")
    plt.title("Usuarios por Año de Nacimiento", pad=20)
    plt.xticks(range(min(anios), max(anios)+1,1), rotation = 90)

   

    # Agregar etiquetas en las barras
    for anio, cantidad in zip(anios, cantidades):
        plt.text(anio, cantidad, str(cantidad), ha='center', va='bottom')

    # Guardar la imagen en un buffer
    buffer = BytesIO()
    plt.savefig(buffer, format="png")
    buffer.seek(0)
    plt.close()

    # Convertir la imagen a base64
    imagen_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
    return imagen_base64

def generar_grafico_respuestas_por_dia():
    with connection.cursor() as cursor:
        cursor.execute(
            "SELECT DATE(Fecha_Ingreso), COUNT(*) FROM botApp_usuario GROUP BY DATE(Fecha_Ingreso)"
        )
        resultados = cursor.fetchall()

    fechas = []
    cantidades = []

    for resultado in resultados:
        fecha, cantidad = resultado
        fechas.append(datetime.strftime(fecha, "%d-%m-%Y"))
        cantidades.append(cantidad)

    plt.plot(fechas, cantidades, marker="o", linestyle="-", color="blue")
    plt.xlabel("Fecha de Respuesta")
    plt.ylabel("Número de Respuestas")
    plt.title("Respuestas por Día", pad=20)
    plt.xticks(rotation = 90)
    plt.tight_layout() 
    
    

    # Agregar los valores de cada punto
    for fecha, cantidad in zip(fechas, cantidades):
        plt.annotate(f"{cantidad}", (fecha, cantidad), textcoords="offset points", xytext=(0,10), ha='center')

    buffer = BytesIO()
    plt.savefig(buffer, format="png")
    buffer.seek(0)
    plt.close()

    imagen_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
    return imagen_base64

def generar_grafico_personas_por_genero():
    with connection.cursor() as cursor:
        cursor.execute(
            "SELECT Genero_Usuario_id, COUNT(*) FROM botApp_usuario GROUP BY Genero_Usuario_id"
        )
        resultados = cursor.fetchall()

    generos = []
    cantidades = []

    for resultado in resultados:
        genero_id, cantidad = resultado
        genero = Genero.objects.get(id=genero_id)
        generos.append(genero.OPC_Genero)
        cantidades.append(cantidad)

    # Crear gráfico de barras con diferentes colores para cada barra
    colores = {'Masculino': 'blue', 'Femenino': 'pink', 'Otro': 'green'}
    plt.bar(generos, cantidades, color=[colores[genero] for genero in generos])

    # Agregar los valores de cada barra
    for i in range(len(generos)):
        plt.text(i, cantidades[i], str(cantidades[i]), ha='center', va='bottom')

    plt.xlabel("Género")
    plt.ylabel("Número de Personas")
    plt.title("Ingresos por Género", pad=20)

    buffer = BytesIO()
    plt.savefig(buffer, format="png")
    buffer.seek(0)
    plt.close()

    imagen_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
    return imagen_base64
    
def generar_grafico_ingresos_por_comuna():
    with connection.cursor() as cursor:
        cursor.execute(
            "SELECT c.Nombre_Comuna, COUNT(*) AS TotalIngresos "
            "FROM botApp_usuario u "
            "JOIN botApp_comuna c ON u.Comuna_Usuario_id = c.id "
            "GROUP BY c.Nombre_Comuna"
        )
        resultados = cursor.fetchall()

    comunas = [result[0] for result in resultados]
    total_ingresos = [result[1] for result in resultados]

    # Configurar el gráfico circular
    fig, ax = plt.subplots()
    wedges, texts, autotexts = ax.pie(total_ingresos, labels=comunas, autopct=lambda pct: f"{pct:.1f}%\n{int(pct/100 * sum(total_ingresos))} ingresos", startangle=90)
    ax.axis('equal')  # Asegura que el gráfico sea un círculo en lugar de una elipse
    ax.set_title('Distribución de Ingresos por Comuna', pad=20)

    # Ajustar el tamaño de la fuente en los textos
    for text, autotext in zip(texts, autotexts):
        text.set(size=8)
        autotext.set(size=8)

    # Convertir el gráfico a base64
    buffer = BytesIO()
    plt.savefig(buffer, format='png')
    buffer.seek(0)
    plt.close()
    imagen_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")

    return imagen_base64

def generar_grafico_referencias():
    with connection.cursor() as cursor:
        cursor.execute(
            "SELECT u.Referencia, COUNT(*) AS TotalIngresos "
            "FROM botApp_usuario u "
            "GROUP BY u.Referencia"
        )
        resultados = cursor.fetchall()

    referencias = [result[0] for result in resultados]
    total_ingresos = [result[1] for result in resultados]

    # Configurar el gráfico circular
    fig, ax = plt.subplots()
    wedges, texts, autotexts = ax.pie(total_ingresos, labels=referencias, autopct=lambda pct: f"{pct:.1f}%\n{int(pct/100 * sum(total_ingresos))} ingresos", startangle=90)
    ax.axis('equal')  # Asegura que el gráfico sea un círculo en lugar de una elipse
    ax.set_title('Distribución de Ingresos por Referencia', pad=20)

    # Ajustar el tamaño de la fuente en los textos
    for text, autotext in zip(texts, autotexts):
        text.set(size=8)
        autotext.set(size=8)

    # Convertir el gráfico a base64
    buffer = BytesIO()
    plt.savefig(buffer, format='png')
    buffer.seek(0)
    plt.close()
    imagen_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")

    return imagen_base64

def generar_grafico_pregunta1():
    with connection.cursor() as cursor:
        cursor.execute(
            "SELECT id_opc_respuesta_id, COUNT(*) FROM botApp_usuariorespuesta WHERE id_opc_respuesta_id IN (8, 9) GROUP BY id_opc_respuesta_id"
        )
        resultados = cursor.fetchall()

    labels = []
    sizes = []
    counts = []

    for resultado in resultados:
        id_opc_respuesta, cantidad = resultado
        opcion_respuesta = PreguntaOpcionRespuesta.objects.get(id=id_opc_respuesta)
        labels.append(opcion_respuesta.OPC_Respuesta)
        sizes.append(cantidad)
        counts.append(f"{opcion_respuesta.OPC_Respuesta} - {cantidad}")

    # Configurar el gráfico circular
    fig, ax = plt.subplots()
    wedges, texts, autotexts = ax.pie(sizes, labels=None, autopct='%1.1f%%', startangle=90, colors=['lightgreen', 'lightcoral', 'lightblue'])
    
    # Configurar las etiquetas del gráfico
    ax.legend(wedges, counts, title="Respuestas", loc="center left", bbox_to_anchor=(1, 0, 0.5, 1))
    
    # Mostrar el gráfico
    plt.title('¿Te has realizado una mamografía?', pad=20)

    # Guardar la imagen en un buffer
    buffer = BytesIO()
    plt.savefig(buffer, format="png", bbox_inches='tight')
    buffer.seek(0)
    plt.close()

    # Convertir la imagen a base64
    imagen_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
    return imagen_base64

def generar_grafico_pregunta2():
    with connection.cursor() as cursor:
        cursor.execute(
            "SELECT id_opc_respuesta_id, COUNT(*) FROM botApp_usuariorespuesta WHERE id_opc_respuesta_id IN (10, 11, 12) GROUP BY id_opc_respuesta_id"
        )
        resultados = cursor.fetchall()

    labels = []
    sizes = []
    counts = []

    for resultado in resultados:
        id_opc_respuesta, cantidad = resultado
        opcion_respuesta = PreguntaOpcionRespuesta.objects.get(id=id_opc_respuesta)
        labels.append(opcion_respuesta.OPC_Respuesta)
        sizes.append(cantidad)
        counts.append(f"{opcion_respuesta.OPC_Respuesta} - {cantidad}")

    # Configurar el gráfico circular
    fig, ax = plt.subplots()
    wedges, texts, autotexts = ax.pie(sizes, labels=None, autopct='%1.1f%%', startangle=90, colors=['lightgreen', 'lightcoral','lightblue' ])
    
    # Configurar las etiquetas del gráfico
    ax.legend(wedges, counts, title="Respuestas", loc="center left", bbox_to_anchor=(1, 0, 0.5, 1))
    
    # Mostrar el gráfico
    plt.title('¿Recuerdas cuando fue tu última mamografía?', pad=20)

    # Guardar la imagen en un buffer
    buffer = BytesIO()
    plt.savefig(buffer, format="png", bbox_inches='tight')
    buffer.seek(0)
    plt.close()

    # Convertir la imagen a base64
    imagen_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
    return imagen_base64

def generar_grafico_pregunta3():
    with connection.cursor() as cursor:
        cursor.execute(
            "SELECT id_opc_respuesta_id, COUNT(*) FROM botApp_usuariorespuesta WHERE id_opc_respuesta_id IN (13, 14, 15, 16) GROUP BY id_opc_respuesta_id"
        )
        resultados = cursor.fetchall()

    labels = []
    sizes = []
    counts = []

    for resultado in resultados:
        id_opc_respuesta, cantidad = resultado
        opcion_respuesta = PreguntaOpcionRespuesta.objects.get(id=id_opc_respuesta)
        labels.append(opcion_respuesta.OPC_Respuesta)
        sizes.append(cantidad)
        counts.append(f"{opcion_respuesta.OPC_Respuesta} - {cantidad}")

    # Configurar el gráfico circular
    fig, ax = plt.subplots()
    wedges, texts, autotexts = ax.pie(sizes, labels=None, autopct='%1.1f%%', startangle=90, colors=['lightgreen', 'lightcoral', 'lightblue', 'lightyellow'])
    
    # Configurar las etiquetas del gráfico
    ax.legend(wedges, counts, title="Respuestas", loc="center left", bbox_to_anchor=(1, 0, 0.5, 1))
    
    # Mostrar el gráfico
    plt.title('Fecha de la última mamografía', pad=20)

    # Guardar la imagen en un buffer
    buffer = BytesIO()
    plt.savefig(buffer, format="png", bbox_inches='tight')
    buffer.seek(0)
    plt.close()

    # Convertir la imagen a base64
    imagen_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
    return imagen_base64

def generar_grafico_pregunta4():
    with connection.cursor() as cursor:
        cursor.execute(
            "SELECT id_opc_respuesta_id, COUNT(*) FROM botApp_usuariorespuesta WHERE id_opc_respuesta_id IN (17, 18, 19, 20) GROUP BY id_opc_respuesta_id"
        )
        resultados = cursor.fetchall()

    labels = []
    sizes = []
    counts = []

    for resultado in resultados:
        id_opc_respuesta, cantidad = resultado
        opcion_respuesta = PreguntaOpcionRespuesta.objects.get(id=id_opc_respuesta)
        labels.append(opcion_respuesta.OPC_Respuesta)
        sizes.append(cantidad)
        counts.append(f"{opcion_respuesta.OPC_Respuesta} - {cantidad}")

    # Configurar el gráfico circular
    fig, ax = plt.subplots()
    wedges, texts, autotexts = ax.pie(sizes, labels=None, autopct='%1.1f%%', startangle=90, colors=['lightgreen', 'lightcoral', 'lightblue', 'lightyellow'])
    
    # Configurar las etiquetas del gráfico
    ax.legend(wedges, counts, title="Respuestas", loc="center left", bbox_to_anchor=(1, 0, 0.5, 1))
    
    # Mostrar el gráfico
    plt.title('¿Tienes los archivos e informe de tu última mamografía?', pad=20)

    # Guardar la imagen en un buffer
    buffer = BytesIO()
    plt.savefig(buffer, format="png", bbox_inches='tight')
    buffer.seek(0)
    plt.close()

    # Convertir la imagen a base64
    imagen_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
    return imagen_base64

def generar_grafico_pregunta5():
    with connection.cursor() as cursor:
        cursor.execute(
            "SELECT id_opc_respuesta_id, COUNT(*) FROM botApp_usuariorespuesta WHERE id_opc_respuesta_id IN (21, 22, 23) GROUP BY id_opc_respuesta_id"
        )
        resultados = cursor.fetchall()

    labels = []
    sizes = []
    counts = []

    for resultado in resultados:
        id_opc_respuesta, cantidad = resultado
        opcion_respuesta = PreguntaOpcionRespuesta.objects.get(id=id_opc_respuesta)
        labels.append(opcion_respuesta.OPC_Respuesta)
        sizes.append(cantidad)
        counts.append(f"{opcion_respuesta.OPC_Respuesta} - {cantidad}")

    # Configurar el gráfico circular
    fig, ax = plt.subplots()
    wedges, texts, autotexts = ax.pie(sizes, labels=None, autopct='%1.1f%%', startangle=90, colors=['lightgreen', 'lightcoral', 'lightblue'])
    
    # Configurar las etiquetas del gráfico
    ax.legend(wedges, counts, title="Respuestas", loc="center left", bbox_to_anchor=(1, 0, 0.5, 1))
    
    # Mostrar el gráfico
    plt.title('¿Te gustaría recibir más información sobre el cuidado y prevención del cáncer de mama?', pad=20)

    # Guardar la imagen en un buffer
    buffer = BytesIO()
    plt.savefig(buffer, format="png", bbox_inches='tight')
    buffer.seek(0)
    plt.close()

    # Convertir la imagen a base64
    imagen_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
    return imagen_base64

def generar_grafico_pregunta6():
    with connection.cursor() as cursor:
        cursor.execute(
            "SELECT id_opc_respuesta_id, COUNT(*) FROM botApp_usuariorespuesta WHERE id_opc_respuesta_id IN (25, 26, 27) GROUP BY id_opc_respuesta_id"
        )
        resultados = cursor.fetchall()

    labels = []
    sizes = []
    counts = []

    for resultado in resultados:
        id_opc_respuesta, cantidad = resultado
        opcion_respuesta = PreguntaOpcionRespuesta.objects.get(id=id_opc_respuesta)
        labels.append(opcion_respuesta.OPC_Respuesta)
        sizes.append(cantidad)
        counts.append(f"{opcion_respuesta.OPC_Respuesta} - {cantidad}")

    # Configurar el gráfico circular
    fig, ax = plt.subplots()
    wedges, texts, autotexts = ax.pie(sizes, labels=None, autopct='%1.1f%%', startangle=90, colors=['lightgreen', 'lightcoral', 'lightblue'])
    
    # Configurar las etiquetas del gráfico
    ax.legend(wedges, counts, title="Respuestas", loc="center left", bbox_to_anchor=(1, 0, 0.5, 1))
    
    # Mostrar el gráfico
    plt.title('¿Tienes un familiar directo con cáncer de mama? (hermana, mama, tía, abuela)', pad=20)

    # Guardar la imagen en un buffer
    buffer = BytesIO()
    plt.savefig(buffer, format="png", bbox_inches='tight')
    buffer.seek(0)
    plt.close()

    # Convertir la imagen a base64
    imagen_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
    return imagen_base64

def generar_grafico_mamografia_si_por_edad():

    with connection.cursor() as cursor:
        cursor.execute(
            """
            SELECT us.edad, COUNT(*) as Cantidad 
            FROM botApp_usuariorespuesta ur JOIN botApp_usuario us ON ur.Rut = us.Rut
            WHERE id_opc_respuesta_id IN (8)
            GROUP BY edad ORDER BY edad ASC
            """


        )
        resultados = cursor.fetchall()

    edades = []
    cantidades = []

    for resultado in resultados:
        edad, cantidad = resultado
        edades.append(edad)
        cantidades.append(cantidad)

    plt.figure(figsize=[13,5])
    plt.bar(edades, cantidades, color="blue")
    plt.xlabel("Edad")
    plt.ylabel("Número de Usuarias")
    plt.title("Mamografías por edad Respuesta Si", pad=20)
    plt.xticks(range(min(edades), max(edades) + 1, 1))
    plt.yticks(range(0,11,1))

    
    # Agregar etiquetas en las barras
    for edad, cantidad in zip(edades, cantidades):
        plt.text(edad, cantidad, str(cantidad), ha='center', va='bottom')

    # Guardar la imagen en un buffer
    buffer = BytesIO()
    plt.savefig(buffer, format="png")
    buffer.seek(0)
    plt.close()

    # Convertir la imagen a base64
    imagen_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
    return imagen_base64

def generar_grafico_mamografia_no_por_edad():

    with connection.cursor() as cursor:
        cursor.execute(
            """
            SELECT us.edad, COUNT(distinct ur.fecha_respuesta) as Cantidad 
            FROM botApp_usuariorespuesta ur JOIN botApp_usuario us ON ur.Rut = us.Rut
            WHERE id_opc_respuesta_id IN (9)
            GROUP BY edad ORDER BY edad ASC
            """


        )
        resultados = cursor.fetchall()

    edades = []
    cantidades = []

    for resultado in resultados:
        edad, cantidad = resultado
        edades.append(edad)
        cantidades.append(cantidad)

    plt.figure(figsize=[13,5])
    plt.bar(edades, cantidades, color="#EFB0C9")
    plt.xlabel("Edad")
    plt.ylabel("Número de Usuarias")
    plt.title("Mamografías por edad Respuesta No", pad=20)
    plt.xticks(range(min(edades), max(edades) + 1, 1))

    
    # Agregar etiquetas en las barras
    for edad, cantidad in zip(edades, cantidades):
        plt.text(edad, cantidad, str(cantidad), ha='center', va='bottom')

    # Guardar la imagen en un buffer
    buffer = BytesIO()
    plt.savefig(buffer, format="png")
    buffer.seek(0)
    plt.close()

    # Convertir la imagen a base64
    imagen_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
    return imagen_base64

def mamografia_por_edad_si_no():
    with connection.cursor() as cursor:
        cursor.execute(
            """
            SELECT us.edad, COUNT(DISTINCT ur.fecha_respuesta) as Cantidad, ur.id_opc_respuesta_id
            FROM botApp_usuariorespuesta ur 
            JOIN botApp_usuario us ON ur.Rut = us.Rut
            WHERE id_opc_respuesta_id IN (8,9)
            GROUP BY edad, ur.id_opc_respuesta_id 
            ORDER BY edad ASC
            """
        )
        resultados = cursor.fetchall()

    edades = []
    cantidades_si = []
    cantidades_no = []

    # Iteramos sobre los resultados
    for resultado in resultados:
        edad, cantidad, respuesta = resultado

        # Si la edad no está en la lista, la agregamos con inicialización de cantidades
        if edad not in edades:
            edades.append(edad)
            cantidades_si.append(0)  
            cantidades_no.append(0)  

        # Obtenemos el índice correspondiente a la edad
        index = edades.index(edad)

        
        if respuesta == 8:
            cantidades_si[index] += cantidad
        elif respuesta == 9: 
            cantidades_no[index] += cantidad

    # Crear el gráfico
    plt.figure(figsize=[18, 8])
    plt.bar(edades, cantidades_si, color="#79addc", label="Cantidad Sí")
    plt.bar(edades, cantidades_no, color="#EFB0C9", bottom=cantidades_si, label="Cantidad No")
    plt.xlabel("Edad")
    plt.ylabel("Número de Usuarias")
    plt.title("Mamografías por Edad", pad=20)
    plt.xticks(range(min(edades), max(edades) + 1, 1))  # Ajuste dinámico del rango de edades
    plt.legend()

    # Agregar etiquetas para las barras de cantidades_si
    for edad, cantidad_si, cantidad_no in zip(edades, cantidades_si, cantidades_no):
        if cantidad_si > 0:
            plt.text(edad, cantidad_si - cantidad_si / 2,  
                 str(cantidad_si), ha='center', va='bottom', color='black')

    # Agregar etiquetas para las barras de cantidades_no
    for edad, cantidad_si, cantidad_no in zip(edades, cantidades_si, cantidades_no):
        if cantidad_no > 0:
            plt.text(edad, cantidad_si + cantidad_no - cantidad_no / 2,  
                str(cantidad_no), ha='center', va='top', color='black')

   # Guardar la imagen en un buffer
    buffer = BytesIO()
    plt.savefig(buffer, format="png")
    buffer.seek(0)
    plt.close()

    # Convertir la imagen a base64
    imagen_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
    return imagen_base64

def generar_grafico_tiempo_trascurrido():

    with connection.cursor() as cursor:
        cursor.execute(
            """
            SELECT tiempo_transc_ult_mamografia, COUNT(*) 
            FROM botApp_ultima_mamografia_anio  
            GROUP BY tiempo_transc_ult_mamografia ORDER BY tiempo_transc_ult_mamografia ASC 
            """
        )
        resultados = cursor.fetchall()

    opciones_anios = ["1", "2", "Más de 2"]
    cantidades = [0, 0, 0]

    for resultado in resultados:
        anio, cantidad = resultado
        if anio == 1:
            cantidades[0] += cantidad
        elif anio == 2:
            cantidades[1] += cantidad
        elif anio > 2:
            cantidades[2] += cantidad

    plt.figure(figsize=[13,5])
    plt.bar(opciones_anios, cantidades, color="blue")
    plt.xlabel("Años transcurridos")
    plt.ylabel("Cantidad de usuarias")
    plt.title("Tiempo transcurrido desde última mamografía", pad=20)

    # Agregar etiquetas en las barras
    for i, cantidad in enumerate(cantidades):
        plt.text(i, cantidad, str(cantidad), ha='center', va='bottom')

    # Guardar la imagen en un buffer
    buffer = BytesIO()
    plt.savefig(buffer, format="png")
    buffer.seek(0)
    plt.close()

    # Convertir la imagen a base64
    imagen_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
    return imagen_base64


def generar_grafico_por_rango_edad():

    with connection.cursor() as cursor:
        cursor.execute(
            """
            SELECT edad, COUNT(*) 
            FROM botApp_usuario 
            GROUP BY edad ORDER BY edad
            """
        )
        resultados = cursor.fetchall()

    opciones_edad = ["Menor de 50 años", "Entre 50 y 69 años", "Mayor a 69 años"]
    cantidades = [0, 0, 0]

    for resultado in resultados:
        edad, cantidad = resultado
        if edad <50:
            cantidades[0] += cantidad
        elif edad >=50 and edad <= 69:
            cantidades[1] += cantidad
        elif edad > 69 :
            cantidades[2] += cantidad

    plt.figure(figsize=[10,5])
    plt.bar(opciones_edad, cantidades, color="#79addc")
    plt.xlabel("Rangos de edad")
    plt.ylabel("Cantidad de usuarias")
    plt.title("Cantidad de usuarias por rango de edad", pad=20)

    # Agregar etiquetas en las barras
    for i, cantidad in enumerate(cantidades):
        plt.text(i, cantidad, str(cantidad), ha='center', va='bottom')



    # Guardar la imagen en un buffer
    buffer = BytesIO()
    plt.savefig(buffer, format="png")
    buffer.seek(0)
    plt.close()

    # Convertir la imagen a base64
    imagen_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
    return imagen_base64

@login_required
def reportes(request):
    data = {
        "imagen_base64_edad": generar_grafico_usuario_por_edad(),
        "imagen_base64_ingresos": generar_grafico_respuestas_por_dia(),
        "imagen_base64_genero":  generar_grafico_personas_por_genero(),
        "imagen_base64_ingresos_comuna": generar_grafico_ingresos_por_comuna(),
        "imagen_base64_pregunta1": generar_grafico_pregunta1(),
        "imagen_base64_pregunta2": generar_grafico_pregunta2(),
        "imagen_base64_pregunta3": generar_grafico_pregunta3(),
        "imagen_base64_pregunta4": generar_grafico_pregunta4(),
        "imagen_base64_pregunta5": generar_grafico_pregunta5(),
        "imagen_base64_pregunta6": generar_grafico_pregunta6(),  
        "imagen_base64_referencias": generar_grafico_referencias(), 
        "imagen_base64_anios_nacimiento": generar_grafico_anio_nacimiento(),
        "imagen_base64_mamografia_si_por_edad":generar_grafico_mamografia_si_por_edad(),
        "imagen_base64_mamografia_no_por_edad":generar_grafico_mamografia_no_por_edad(),
        "imagen_base64_mamografia_por_edad_si_no": mamografia_por_edad_si_no(),
        "imagen_base64_tiempo_transc": generar_grafico_tiempo_trascurrido(),
        "imagen_base64_rango_edad": generar_grafico_por_rango_edad(),

            }
    return render(request, "reportes.html", data)


# --------------------- Formulario WEB --------------------- #
@login_required
def formulario(request):
    data = {
        "formUsuario": UsuarioForm(),
        "preguntas": Pregunta.objects.all(),
        "usuarios": User.objects.all(),
    }

    if request.method == "POST":
        form_usuario = UsuarioForm(request.POST)

        if form_usuario.is_valid():
            form_usuario.instance.id_usuario = request.POST.get("id_usuario")
            usuario = form_usuario.save()

            for pregunta in data["preguntas"]:
                respuesta = request.POST.get(f"pregunta_{pregunta.id}")
                opc_respuesta = OPC_Respuesta(
                    id_pregunta=pregunta, OPC_Respuesta=respuesta
                )
                opc_respuesta.save()
                respuesta_usuario = RespuestaUsuario(
                    id_usuario=usuario,
                    id_pregunta=pregunta,
                    id_opc_respuesta=opc_respuesta,
                )
                respuesta_usuario.save()

            messages.success(request, "Datos guardados correctamente")
            form_usuario = UsuarioForm()
            return redirect(to="home")

        else:
            print(form_usuario.errors)
            messages.error(
                request,
                "La persona debe tener más de 18 años y haber nacido después de 1930.",
            )
            form_usuario = UsuarioForm()

    return render(request, "formulario.html", data)


# --------------------- Preguntas --------------------- #

# Listar Preguntas
@login_required
def listarPreguntas(request):
    Preguntas = Pregunta.objects.all()
    data = {
        "preguntas": Preguntas,
    }
    return render(request, "preguntas/listarPreguntas.html", data)


# Modificar Pregunta
@login_required
def modificarPregunta(request, id):
    Preguntas = Pregunta.objects.get(id=id)
    data = {"form": PreguntaForm(instance=Preguntas)}

    if request.method == "POST":
        formulario = PreguntaForm(
            data=request.POST, instance=Preguntas, files=request.FILES
        )
        if formulario.is_valid():
            formulario.save()
            messages.success(request, "Modificado Correctamente")
            return redirect(to="listarPreguntas")
        data["form"] = formulario

    return render(request, "preguntas/modificarPreguntas.html", data)


# Eliminar Pregunta
@login_required
def eliminarPregunta(request, id):
    Preguntas = Pregunta.objects.get(id=id)
    Preguntas.delete()
    messages.success(request, "Eliminado Correctamente")
    return redirect(to="listarPreguntas")


# Crear Pregunta
@login_required
def crearPregunta(request):
    data = {"form": PreguntaForm()}

    if request.method == "POST":
        formulario = PreguntaForm(data=request.POST, files=request.FILES)
        if formulario.is_valid():
            formulario.save()
            messages.success(request, "Pregunta Creada Correctamente")
        else:
            data["form"] = formulario

    return render(request, "preguntas/crearPreguntas.html", data)


# --------------------- Mensajes --------------------- #
@login_required
def homeMensajes(request):
    Mensajes = MensajeContenido.objects.all()
    data = {
        "mensajes": Mensajes,
    }
    return render(request, "mensajes/homeMensajes.html", data)

@login_required
def crearMensaje(request):
    data = {"form": MensajeContenidoForm()}

    if request.method == "POST":
        formulario = MensajeContenidoForm(data=request.POST, files=request.FILES)
        if formulario.is_valid():
            formulario.save()
            messages.success(request, "Mensaje Creado Correctamente")
        else:
            data["form"] = formulario

    return render(request, "mensajes/crearMensajes.html", data)

@login_required
def modificarMensaje(request, id):
    Mensajes = MensajeContenido.objects.get(id=id)
    data = {"form": MensajeContenidoForm(instance=Mensajes)}

    if request.method == "POST":
        formulario = MensajeContenidoForm(
            data=request.POST, instance=Mensajes, files=request.FILES
        )
        if formulario.is_valid():
            formulario.save()
            messages.success(request, "Modificado Correctamente")
            return redirect(to="mensajesHome")
        data["form"] = formulario

    return render(request, "mensajes/modificarMensajes.html", data)

@login_required
def eliminarMensaje(request, id):
    Mensajes = MensajeContenido.objects.get(id=id)
    Mensajes.delete()
    messages.success(request, "Eliminado Correctamente")
    return redirect(to="mensajesHome")

# --------------------- Api --------------------- #

@login_required
def apiHome(request):
    return render(request, "api/apiHome.html")

#Usuario
class UsuarioViewSet(viewsets.ModelViewSet):
    queryset = Usuario.objects.all()
    serializer_class = UsuarioSerializer

#RespuestaUsuario
class UsuarioRespuestaViewSet(viewsets.ModelViewSet):
    queryset = UsuarioRespuesta.objects.all()
    serializer_class = UsuarioRespuestaSerializer

#TextoPreguntaUsuario
class UsuarioTextoPreguntaViewSet(viewsets.ModelViewSet):
    queryset = UsuarioTextoPregunta.objects.all()
    serializer_class = UsuarioTextoPreguntaSerializer
    
#MensajeContenido
class MensajeContenidoViewSet(viewsets.ModelViewSet):
    queryset = MensajeContenido.objects.all()
    serializer_class = MensajeContenidoSerializer

class UsuarioRespuestaAPIView(APIView):
    authentication_classes = [SessionAuthentication]
    permission_classes = [IsAdminUser]
    def get(self, request):
        respuestas = UsuarioRespuesta.objects.all()
        serializer = UsuarioRespuestaSerializer(respuestas, many=True)
        return Response(serializer.data)

class UsuarioTextoPreguntaAPIView(APIView):
    authentication_classes = [SessionAuthentication]
    permission_classes = [IsAdminUser]
    def get(self, request):
        textos_pregunta = UsuarioTextoPregunta.objects.all()
        serializer = UsuarioTextoPreguntaSerializer(textos_pregunta, many=True)
        return Response(serializer.data)

class UsuarioAPIView(APIView):
    authentication_classes = [SessionAuthentication]
    permission_classes = [IsAdminUser]
    def get(self, request):
        usuarios = Usuario.objects.all()
        serializer = UsuarioSerializer(usuarios, many=True)
        return Response(serializer.data)
    
class MensajeContenidoAPIView(APIView):
    authentication_classes = [SessionAuthentication]
    permission_classes = [IsAdminUser]
    def get(self, request):
        mensajes = MensajeContenido.objects.all()
        serializer = MensajeContenidoSerializer(mensajes, many=True)
        return Response(serializer.data)
    
class ObtenerID(APIView):
    def get(self, request):
        # Obtener la fecha de hoy
        fecha_hoy = date.today()
        
        # Buscar un registro en la tabla que coincida con la fecha de hoy
        registro_hoy = MensajeContenido.objects.filter(fecha=fecha_hoy).first()
        
        if registro_hoy:
            # Si se encuentra un registro para la fecha de hoy, devolverlo
            return Response({'id': registro_hoy.id, 'texto': registro_hoy.texto, 'genero': registro_hoy.Genero_Usuario.OPC_Genero})
        else:
            # Si no se encuentra ningún registro para la fecha de hoy, devolver un código de error (por ejemplo, "1")
            return Response({'error_code': '1'})
# --------------------- Api --------------------- #